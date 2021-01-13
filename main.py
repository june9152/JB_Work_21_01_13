


import pandas as pd
from openpyxl import load_workbook
import os




SheetName = 'RM'
ExcelPath = r'E:/'
FinalPath = r'E:/' + SheetName + '/'
load_wb = load_workbook(ExcelPath+'2020_손정빈_세금계산서_12월.xlsx',data_only=True)
load_ws = load_wb[SheetName]

def Data_Convert(str):
    str=str.replace('2020',"20")
    str=str.replace('-',"")
    str=str.replace(' 00:00:00',"")
    return str
def createFolder(directory):
    try:
        if not os.path.exists(directory):
            os.makedirs(directory)
    except OSError:
        print('Error: Creating Dir' + directory)




def Excel_RowLength(ws):
    i = 3
    while True:
        name = ws['D'+ str(i)].value
        # print(name)
        if name == None:
            break
        else:
            foldername = name
            # print(foldername)
        i = i + 1
    return i

def Excel_RowLengthFunc(ws):
    Excel_Length = Excel_RowLength(ws)
    # print(Excel_Length)
    for i in range(3,Excel_Length):
        name = ws['D'+ str(i)].value
        ret = JB_Work(name)
        print(ret)
    return ret 

def JB_Work(str):
    data = '기자재'
    data2 = '수선비'
    data3 = '시약'
    data4 = '안전'
    data5 = '특허'
    data6 = '시험의뢰비'
    if data in str:
        ret = 'B'
    elif data2 in str:
        ret = 'B'
    elif data3 in str:
        ret = 'A' 
    elif data4 in str:
        ret = 'E'      
    elif data5 in str:
        ret = 'E'    
    elif data6 in str:
        ret = 'C'      
    else:
        ret = ''


    return ret




    



def JB_FolderCreate():
    Excel_Length = Excel_RowLength(load_ws)

    for i in range(3,Excel_Length):
        # foldername = load_ws['D'+ str(i)].value
        foldername = JB_Work(str(load_ws['D'+ str(i)].value))
        foldername += '-'
        foldername += str(load_ws['B'+ str(i)].value)
        foldername += '-'
        foldername += str(load_ws['C'+ str(i)].value)
        foldername += '-'
        foldername += Data_Convert(str(load_ws['F'+ str(i)].value))
        foldername += '-'
        foldername += str(load_ws['E'+ str(i)].value)
        foldername += '-'
        foldername += str(format(load_ws['G'+ str(i)].value,","))
        createFolder(FinalPath+foldername)
        print(foldername)    

# Excel_RowLengthFunc(load_ws)


JB_FolderCreate()



# print(load_ws['B2'.value])
#  print(filedata)




